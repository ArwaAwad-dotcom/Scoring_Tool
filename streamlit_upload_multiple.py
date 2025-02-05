# -*- coding: utf-8 -*-
"""
Created on Fri Jan 24 13:16:22 2025

@author: user
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from analysis import get_relevant_data
from analysis import calculate_weights
from openpyxl import load_workbook
from analysis import combine_baseline_data
from analysis import calculate_psg_score
from analysis import calculate_psg_score_v2
from analysis import level_one_aggregation
import numpy as np


def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    processed_data = output.getvalue()
    return processed_data




def process_files(uploaded_files,level_one_list,level_two_list):
    combined_data = pd.DataFrame()

    for file in uploaded_files:

        # Perform some basic analysis (e.g., add a column with filename)
        df=get_relevant_data(file,level_one_list,level_two_list)
        columns_numeric=df.columns
        last_column=len(columns_numeric)-4
        columns_numeric=columns_numeric[2:last_column]
        for cl in columns_numeric: 
            df[cl]=pd.to_numeric(df[cl], errors='coerce')
            
            
        combined_data = pd.concat([combined_data, df], ignore_index=True)

    return combined_data



def convert_df_to_excel(df):
    """Converts a DataFrame to an Excel file and returns it as a downloadable object."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Processed_Data')
    processed_file = output.getvalue()
    return processed_file


def save_workbook_to_bytes(workbook):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)  # Reset the pointer to the beginning of the stream
    return output


ctf='Mindsets PMS_Comprehensive Talent Assessment Form_v03.xlsx'

file_psg='PSG_Level_Matrix.xlsx'

#Read all level one
level_one=pd.read_excel('Level_one.xlsx')
level_one_list=list(level_one['Level 1'])


#Read all level 2
level_two=pd.read_excel('Level_two.xlsx')
level_two_list=list(level_two['Level 2'])

#combine in a dataframe

levels_one_two=pd.DataFrame({'Level 1': level_one_list,'Level 2':level_two_list})



# Streamlit App
st.title("Talent Assessment")

st.write("Upload all the EAF and Baseline Assessments for the talent.")

# Upload multiple files

psg_options = ['PSG 10', 'PSG 11', 'PSG 12-13', 'PSG 14', 'PSG 15-16', 'PSG 17','PSG 18']

# Dropdown for multiple selections
selected_psg = st.selectbox("Select one or more PSGs:", psg_options)

st.write("You selected",selected_psg)


uploaded_files = st.file_uploader("Upload Excel Files", type=['xlsx'], accept_multiple_files=True)

if uploaded_files:
    st.success(f"{len(uploaded_files)} file(s) uploaded successfully!")
    
    # Process the files
    with st.spinner("Processing files..."):
        
        eafs=[]
        bool_bas=False
        
        for name in uploaded_files:
            print(name.name)
            if 'Engagement' in name.name or 'Engegement'  in name.name:
                eafs.append(name)
            elif 'Baselining' in name.name:
                bas=name
                bool_bas=True
                

        
        combined_data = process_files(eafs,level_one_list,level_two_list)
        
        
        print(combined_data.head())
        if  bool_bas:
            combined_data=combine_baseline_data(bas,combined_data,level_one_list,level_two_list)
        

        matrix_data=calculate_weights(combined_data)
        
        all_col=combined_data.columns
        psg_levels=list(all_col[3:10])
        
        psg_final_grade=calculate_psg_score_v2(matrix_data,levels_one_two,psg_levels,file_psg)
        
        
        
        
        workbook = load_workbook(ctf)
        sheet = workbook["Assessment"]
        
        
        start_row=6
        start_col=3
        
        for i, row in enumerate(psg_final_grade):
            sheet.cell(row=start_row , column=start_col+i).value = row
            
        
        

        start_row=7
        start_col=3

        for i, row in enumerate(matrix_data):
            for j, value in enumerate(row):
                sheet.cell(row=start_row + i, column=start_col + j).value = value
        
        workbook_data = save_workbook_to_bytes(workbook)
                



    
    # Display processed data
    st.write("Processed Data:")
    st.dataframe(combined_data)
    bool_bas=False
    
    
    #Get the levels and psg level

    level_two_list=list(level_two['Level 2'])
    unique_engagement_number=list(combined_data['Engagement Name'].unique())

    psg_levels=selected_psg



    
    dict_engagement={}

    all_engagement={"Level 1":['Professional Behavior','Core Business Skills','Management Skills','Technical Knowledge']}
    
    all_col=combined_data.columns
    psg_levels_list=list(all_col[3:10])



    psg_level_code= ['PSG 10', 'PSG 11', 'PSG 12-13', 'PSG 14', 'PSG 15-16', 'PSG 17','PSG 18']
    psg_levels_specific=selected_psg
    index_psg=psg_level_code.index(psg_levels_specific)
    

    for un in unique_engagement_number:
        
        unique_engagement_matrix=np.zeros((len(all_engagement["Level 1"]),7))

        unique_data_frame=combined_data[combined_data['Engagement Name']==un]
        unique_data_frame=unique_data_frame.iloc[:,:10]
        
        for i,lev in enumerate(all_engagement["Level 1"]):
            print(lev)
            unique_data_frame_lev=unique_data_frame[unique_data_frame['Level 1']==lev]
            
            weight_array=np.array(list(unique_data_frame_lev['Weight']))
            weight_array= np.nan_to_num(weight_array, nan=0)

            
            for k,prof in enumerate(psg_levels_list):
                array_prof=np.array(list(unique_data_frame_lev[prof]))
                array_prof= np.nan_to_num(array_prof, nan=0)

                dot_product=np.dot(weight_array, array_prof)
                unique_engagement_matrix[i,k]=dot_product/(np.sum(weight_array))
                
        array_keep=list(unique_engagement_matrix[:,index_psg])
        
        all_engagement[un]= array_keep
        
        if un==unique_engagement_number[-1] and un=="Previous Year":
            unique_engagement_matrix_bas=np.zeros((len(level_two_list),7))
            unique_engagement_matrix_bas_without_multiple=np.zeros((len(level_two_list),7))
            unique_data_frame_bas=combined_data[combined_data['Engagement Name']==un]
            unique_data_frame_bas=unique_data_frame_bas.iloc[:,:10]
            
            
            list_grades=[]
            for i in range(len(unique_data_frame_bas)):
                get_weight=unique_data_frame_bas['Weight'].iloc[i]
                list_grades_one=np.array(list(get_weight*unique_data_frame_bas.iloc[i,3:]))
                unique_engagement_matrix_bas[i]= list_grades_one
                unique_engagement_matrix_bas_without_multiple[i]=unique_data_frame_bas.iloc[i,3:]
                array_baseline=calculate_psg_score_v2(unique_engagement_matrix,levels_one_two,psg_levels,file_psg)
                array_baseline_without=calculate_psg_score_v2(unique_engagement_matrix_bas_without_multiple,levels_one_two,psg_levels,file_psg)

                baseline=pd.DataFrame({"Levels":psg_options,"Grade Baseline (Weighted_Average)":array_baseline, "Grade Baseline (Without weighted)": array_baseline_without})
                base_data_average =convert_df_to_excel(baseline)
                bool_bas=True
                
            
        
    
    
    data_frame_averages=pd.DataFrame(all_engagement)
    excel_data_average =convert_df_to_excel(data_frame_averages)

    
    
    

    # Allow user to download the processed data
    processed_file = convert_df_to_excel(combined_data)

    st.download_button(
        label="Download CTF",
        data=workbook_data,
        file_name="CTF.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    st.download_button(
        label="Download Engagement Average",
        data=excel_data_average ,
        file_name="eng_avg.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    if  bool_bas:
        st.download_button(
            label="Baseline Score",
            data=base_data_average ,
            file_name="base_avg.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        

